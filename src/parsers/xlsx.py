"""XLSX parser that extracts sheet-wise tabular content with metadata."""

from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover - optional dependency
    openpyxl = None

from .base import BaseParser, ParserError


class XlsxParser(BaseParser):
    """Parse .xlsx workbooks into normalized text and sheet metadata."""

    source_type = "xlsx"
    supported_extensions = (".xlsx",)

    def parse(self, file_path: str | Path):
        path = Path(file_path)
        if not path.exists():
            raise ParserError(f"XLSX file not found: {path}")
        if openpyxl is None:
            raise ParserError("openpyxl is required to parse .xlsx files")

        workbook = self._load_workbook(path)
        output_lines: list[str] = []
        sheet_metadata: list[dict[str, Any]] = []

        for worksheet in workbook.worksheets:
            rows = self._extract_sheet_rows(worksheet)
            if not rows:
                continue

            header = rows[0]
            data_rows = rows[1:]
            output_lines.append(f"Sheet: {worksheet.title}")
            output_lines.append(" | ".join(header))
            output_lines.extend(" | ".join(row) for row in data_rows)
            output_lines.append("")

            sheet_metadata.append(
                {
                    "name": worksheet.title,
                    "range": worksheet.calculate_dimension(),
                    "column_count": len(header),
                    "row_count": len(data_rows),
                }
            )

        output_lines = self._strip_outer_whitespace(output_lines)
        if not output_lines:
            raise ParserError(f"XLSX file is empty: {path}")

        metadata = {
            "format": self.source_type,
            "sheet_count": len(sheet_metadata),
            "sheet_names": [sheet["name"] for sheet in sheet_metadata],
            "sheets": sheet_metadata,
        }
        return self._build_document(
            text="\n".join(output_lines),
            file_path=path,
            source_type=self.source_type,
            metadata=metadata,
        )

    def _load_workbook(self, file_path: Path):
        try:
            return openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as exc:  # pragma: no cover - backend exceptions vary
            raise ParserError(
                f"Unable to open XLSX workbook {file_path}: {exc}"
            ) from exc

    def _extract_sheet_rows(self, worksheet: Any) -> list[list[str]]:
        rows: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            cleaned_row = [self._format_cell(cell) for cell in row]
            if any(cell.strip() for cell in cleaned_row):
                rows.append(cleaned_row)
        return rows

    def _format_cell(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, (date, time)):
            return value.isoformat()
        return str(value).strip()
